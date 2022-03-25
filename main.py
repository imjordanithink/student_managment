from email import header
from itertools import count
from sqlite3 import Row
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Color, PatternFill
from openpyxl.styles import Font
import datetime
from datetime import datetime as dt
import os
import re

def wb_pr_sorter(work_book, class_label): #3 Main parts. 1 - Cell scanning and list/dictionary creation. 2 - Workbook setup and styling. 3 - Ordered appending from dictionaries to worksheet.
    # 1
    wb = load_workbook(work_book)
    ws = wb.active

    sort_dict = {1:"A", 2:"B", 3:"C", 4:"D"}

    a_list = []
    b_list = []
    c_list = []
    d_list = []

    for row in range(3, ws.max_row + 1):
        if ws["D" + str(row)].value == sort_dict[1]:
            for col in range(1, ws.max_column + 1):
                char = get_column_letter(col)
                point = ws[char + str(row)].value
                a_list.append(point)
        elif ws["D" + str(row)].value == sort_dict[2]:
            for col in range(1, ws.max_column + 1):
                char = get_column_letter(col)
                point = ws[char + str(row)].value
                b_list.append(point)
        elif ws["D" + str(row)].value == sort_dict[3]:
            for col in range(1, ws.max_column + 1):
                char = get_column_letter(col)
                point = ws[char + str(row)].value
                c_list.append(point)
        elif ws["D" + str(row)].value == sort_dict[4]:
            for col in range(1, ws.max_column + 1):
                char = get_column_letter(col)
                point = ws[char + str(row)].value
                d_list.append(point)

    a_dict = {a_list[i] : a_list[i+1 : i+5] for i in range(0, len(a_list), 5)}
    b_dict = {b_list[i] : b_list[i+1 : i+5] for i in range(0, len(b_list), 5)}
    c_dict = {c_list[i] : c_list[i+1 : i+5] for i in range(0, len(c_list), 5)}
    d_dict = {d_list[i] : d_list[i+1 : i+5] for i in range(0, len(d_list), 5)}
    
    # 2
    wb.save("{}".format(work_book))
    os.remove("{}".format(work_book))

    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = str(work_book)

    header_list = ["First Name", "Last Name", "Belt", "Priority Rating", "Future Test Date"]
    title_font = Font(size = 40)
    ws2.cell(1, 1).font = title_font
    ws2["A1"] = class_label
    ws2.merge_cells("A1:E1")
    ws2.row_dimensions[1].height = 35

    color_dict = {"Beg":"A6D6D8", "Int":"BB9ED1", "Adv":"FFA07A", "Tee":"78BA97", "Adu":"B57777", "InA":"FFA07A"}
    color = color_dict[class_label[4:8]]
    
    ws2["A1"].fill = PatternFill(patternType="solid", fgColor=color)

    ws2.append(header_list)

    # 3
    for col in range(1, 6):
        char = get_column_letter(col)
        ws2[char + "2"].font = Font(bold=True, underline="single")

    for col in range(1, 6):
        char = get_column_letter(col)
        ws2.column_dimensions[char].width = 18
    
    if len(a_dict) == 0:
        pass
    else:
       for person in a_dict:
            points = list(a_dict[person])
            ws2.append([person] + points)
    if len(b_dict) == 0:
        pass
    else:
       for person in b_dict:
            points = list(b_dict[person])
            ws2.append([person] + points)
    if len(c_dict) == 0:
        pass
    else:
       for person in c_dict:
            points = list(c_dict[person])
            ws2.append([person] + points)
    if len(d_dict) == 0:
        pass
    else:
       for person in d_dict:
            points = list(d_dict[person])
            ws2.append([person] + points)

    wb2.save("{}.xlsx".format(class_label))
    print("\nSheet Creation Successful!\n")
    display()

def wb_belt_sorter(work_book, class_label): #3 Main parts. 1 - Cell scanning and list/dictionary creation. 2 - Workbook setup and styling. 3 - Ordered appending from dictionaries to worksheet.
    # 1
    wb = load_workbook(work_book)
    ws = wb.active

    sort_dict = {1 : "No", 2 : "White", 3:"Yellow", 4:"Orange", 5:"Green", 6:"Blue", 7:"Purple", 8:"Brown", 9:"Red", 10:"Black Stripe", 11:"Black"}

    #dicts will be added to .xlsx in belt order. Probably ineficient but oh well
    no_list = []
    white_list = []
    yellow_list = []
    orange_list = []
    green_list = []
    blue_list = []
    purple_list = []
    brown_list = []
    red_list = []
    black_stripe_list = []
    black_list = []
    
    for row in range(3, ws.max_row + 1):
        if ws["C" + str(row)].value == sort_dict[1]:
            for col in range(1, ws.max_column + 1):
                char = get_column_letter(col)
                point = ws[char + str(row)].value
                no_list.append(point)
        elif ws["C" + str(row)].value == sort_dict[2]:
            for col in range(1, ws.max_column + 1):
                char = get_column_letter(col)
                point = ws[char + str(row)].value
                white_list.append(point)
        elif ws["C" + str(row)].value == sort_dict[3]:
            for col in range(1, ws.max_column + 1):
                char = get_column_letter(col)
                point = ws[char + str(row)].value
                yellow_list.append(point)
        elif ws["C" + str(row)].value == sort_dict[4]:
            for col in range(1, ws.max_column + 1):
                char = get_column_letter(col)
                point = ws[char + str(row)].value
                orange_list.append(point)
        elif ws["C" + str(row)].value == sort_dict[5]:
            for col in range(1, ws.max_column + 1):
                char = get_column_letter(col)
                point = ws[char + str(row)].value
                green_list.append(point)
        elif ws["C" + str(row)].value == sort_dict[6]:
            for col in range(1, ws.max_column + 1):
                char = get_column_letter(col)
                point = ws[char + str(row)].value
                blue_list.append(point)
        elif ws["C" + str(row)].value == sort_dict[7]:
            for col in range(1, ws.max_column + 1):
                char = get_column_letter(col)
                point = ws[char + str(row)].value
                purple_list.append(point)
        elif ws["C" + str(row)].value == "Brown":
            for col in range(1, ws.max_column + 1):
                char = get_column_letter(col)
                point = ws[char + str(row)].value
                brown_list.append(point)
        elif ws["C" + str(row)].value == "Red":
            for col in range(1, ws.max_column + 1):
                char = get_column_letter(col)
                point = ws[char + str(row)].value
                red_list.append(point)
        elif ws["C" + str(row)].value == sort_dict[10]:
            for col in range(1, ws.max_column + 1):
                char = get_column_letter(col)
                point = ws[char + str(row)].value
                black_stripe_list.append(point)
        elif ws["C" + str(row)].value == sort_dict[11]:
            for col in range(1, ws.max_column + 1):
                char = get_column_letter(col)
                point = ws[char + str(row)].value
                black_list.append(point)

    no_dict = {no_list[i] : no_list[i+1 : i+5] for i in range(0, len(no_list), 5)}
    white_dict = {white_list[i] : white_list[i+1 : i+5] for i in range(0, len(white_list), 5)}
    yellow_dict = {yellow_list[i] : yellow_list[i+1 : i+5] for i in range(0, len(yellow_list), 5)}
    orange_dict = {orange_list[i] : orange_list[i+1 : i+5] for i in range(0, len(orange_list), 5)}
    green_dict = {green_list[i] : green_list[i+1 : i+5] for i in range(0, len(green_list), 5)}
    blue_dict = {blue_list[i] : blue_list[i+1 : i+5] for i in range(0, len(blue_list), 5)}
    purple_dict = {purple_list[i] : purple_list[i+1 : i+5] for i in range(0, len(purple_list), 5)}
    brown_dict = {brown_list[i] : brown_list[i+1 : i+5] for i in range(0, len(brown_list), 5)}
    red_dict = {red_list[i] : red_list[i+1 : i+5] for i in range(0, len(red_list), 5)}
    black_stripe_dict = {black_stripe_list[i] : black_stripe_list[i+1 : i+5] for i in range(0, len(black_stripe_list), 5)}
    black_dict = {black_list[i] : black_list[i+1 : i+5] for i in range(0, len(black_list), 5)}

    # 2
    wb.save("{}".format(work_book))
    os.remove("{}.xlsx".format(class_label))

    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = str(work_book)

    header_list = ["First Name", "Last Name", "Belt", "Priority Rating", "Future Test Date"]
    title_font = Font(size = 40)
    ws2.cell(1, 1).font = title_font
    ws2["A1"] = class_label
    ws2.merge_cells("A1:E1")
    ws2.row_dimensions[1].height = 35

    color_dict = {"Beg":"A6D6D8", "Int":"BB9ED1", "Adv":"FFA07A", "Tee":"78BA97", "Adu":"B57777", "InA":"FFA07A"}
    color = color_dict[class_label[4:8]]
    
    ws2["A1"].fill = PatternFill(patternType="solid", fgColor=color)

    ws2.append(header_list)
    
    # 3
    for col in range(1, 6):
        char = get_column_letter(col)
        ws2[char + "2"].font = Font(bold=True, underline="single")

    for col in range(1, 6):
        char = get_column_letter(col)
        ws2.column_dimensions[char].width = 18
   
    if len(no_dict) == 0:
        pass
    else:
       for person in no_dict:
            points = list(no_dict[person])
            ws2.append([person] + points)
    if len(white_dict) == 0:
        pass
    else:
        for person in white_dict:
            points = list(white_dict[person])
            ws2.append([person] + points)
    if len(yellow_dict) == 0:
        pass
    else:
        for person in yellow_dict:
            points = list(yellow_dict[person])
            ws2.append([person] + points)
    if len(orange_dict) == 0:
        pass
    else:
        for person in orange_dict:
            points = list(orange_dict[person])
            ws2.append([person] + points)
    if len(green_dict) == 0:
        pass
    else:
        for person in green_dict:
            points = list(green_dict[person])
            ws2.append([person] + points)
    if len(blue_dict) == 0:
        pass
    else:       
        for person in blue_dict:
            points = list(blue_dict[person])
            ws2.append([person] + points)
    if len(purple_dict) == 0:
        pass
    else:
        for person in purple_dict:
            points = list(purple_dict[person])
            ws2.append([person] + points)
    if len(brown_dict) == 0:
        pass
    else:       
        for person in brown_dict:
            points = list(brown_dict[person])
            ws2.append([person] + points)
    if len(red_dict) == 0:
        pass
    else:
        for person in red_dict:
            points = list(red_dict[person])
            ws2.append([person] + points)
    if len(black_stripe_dict) == 0:
        pass
    else:
        for person in black_stripe_dict:
            points = list(black_stripe_dict[person])
            ws2.append([person] + points)
    if len(black_dict) == 0:
        pass
    else:
        for person in black_dict:
            points = list(black_dict[person])
            ws2.append([person] + points)
    
    wb2.save("{}.xlsx".format(class_label))
    print("\nSheet Creation successful!\n")
    display()

def class_wb_creator(student_dict, class_label, sort):
    wb = Workbook()
    ws = wb.active
    ws.title = class_label

    header_list = ["First Name", "Last Name", "Belt", "Priority Rating", "Future Test Date"]
    title_font = Font(size = 40)
    ws.cell(1, 1).font = title_font
    ws["A1"] = class_label
    ws.merge_cells("A1:E1")
    ws.row_dimensions[1].height = 40

    color_dict = {"Beg":"A6D6D8", "Int":"BB9ED1", "Adv":"FFA07A", "Tee":"78BA97", "Adu":"B57777", "InA":"FFA07A"}
    color = color_dict[class_label[4:8]]
    
    ws["A1"].fill = PatternFill(patternType="solid", fgColor=color)

    ws.append(header_list)

    # Sets title row to be bold and underlined
    for col in range(1, 6):
        char = get_column_letter(col)
        ws[char + "2"].font = Font(bold=True, underline="single")
    # Sets column width
    for col in range(1, 6):
        char = get_column_letter(col)
        ws.column_dimensions[char].width = 18
    
    for person in student_dict:
        points = list(student_dict[person])
        ws.append([person] + points)

    wb.save("{}.xlsx".format(class_label))

    if sort == "b":
        wb_belt_sorter("{}.xlsx".format(class_label), class_label)
    elif sort == "pr":
        wb_pr_sorter("{}.xlsx".format(class_label), class_label)


def wb_data_compiler():
    print("""
    ------- Class Legend ------ | --------- Format Eg's ---------
    Monday = Mon                |
    Tuesday = Tue               |            
    Wednsday = Wed              |            Mon-Adv
    Thursday = Thu              |
    Friday = Fri                |            Tue-Beg
    Saturday = Sat              |            
                                |            Wed-Int
    Beginner = Beg              |
    Intermediate = Int          |            Thu-Adu
    Advanced = Adv              |
    Teen = Tee                  |            Thu-Tee
    Adult = Adu                 |            
                                |            Fri-Int
    **************************  |            
    Int and Adv = InA           |            Sat-Beg
                                |            
    No Class = na               |            Sat-InA

    -----------------------------------------------------------
    """)
    x = input("What day do you want to generate a sheet for?: ")
    if re.search(r"(\w{3})[-](\w{3})", x) == None:
        print("Invalid Entry. Try again\n")
        wb_data_compiler()
    sort = input("Would you like the sheet to be organized by Belt or Priority Rating (b/pr)?: ")
    if sort != "b" and sort != "pr":
        print("Invalid Entry. Try again\n")
        wb_data_compiler()

    class_dict = {
        "Mon":{"Beg":"4:30-5:15pm", "Adv":"6:00-6:45pm", "Int":"6:45-7:30pm"},
        "Tue":{"Int":"4:30-5:15pm", "Adv":"5:15-6:00pm", "Beg":"6:00-6:45pm", "Tee":"6:45-7:30pm", "Adu":"7:30-8:30pm"},
        "Wed":{"Adv":"4:30-5:15pm", "Beg":"5:15-6:00pm", "Int":"6:30-7:15pm"},
        "Thu":{"Adv":"4:30-5:15pm", "Int":"5:15-6:00pm", "Beg":"6:30:7:15pm", "Tee":"6:45-7:30pm", "Adu":"7:30-8:30pm"},
        "Fri":{"Int":"4:30-5:15pm", "Beg":"5:15-6:00pm", "Adv":"6:45-7:30pm", "Adu":"7:30-8:30pm"},
        "Sat":{"Beg":"10:00-10:45am", "InA":"11:15-12:00pm"}
        }

    wbsd = load_workbook("Student_Data.xlsx")
    wssd = wbsd.active

    
    cell_list = []
    for row in range(2, wssd.max_row + 1):
            for col in range(7, wssd.max_column):
                char = get_column_letter(col)
                cell = char + str(row)
                if wssd[cell].value == None:
                    wssd[cell] = "N/A"
                elif x == str(wssd[cell].value):
                    cell_list.append(cell)

    col_list = []
    for col in range(1,5):
        char = get_column_letter(col)
        col_list.append(char)
    row_list = []
    for i in range(len(cell_list)):
        cell = cell_list[i]
        row = cell[1:len(cell)]
        row_list.append(row)
    
    
    first_cells = []
    for row in row_list:
        for col in col_list:
                cell = col + row
                first_cells.append(cell)
    
    insertion_row_list = []
    for i in range(0, len(first_cells), 4):
        cell = first_cells[i]
        row = cell[1:len(cell)]
        insertion_row_list.append(row)

    testing_cells = []
    for row in insertion_row_list:
        cell = "F" + str(row)
        testing_cells.append(cell)
    

    index = 0
    count = 0
    for i in range(len(first_cells) + 1):
        count += 1
        if count == 5:
            first_cells.insert(i, testing_cells[index])
            index+=1
            count = 0

    if len(first_cells) < 1:
        print("Invalid Entry. Try again")
        wb_data_compiler()
    else:
        first_cells.append(testing_cells[-1])

    last_list = []
    for cell in first_cells:
        point = wssd[cell].value
        last_list.append(point)

    student_dict = {last_list[i]: last_list[i +1: i + 5] for i in range(0, len(last_list), 5)}

    wbsd.save("Student_Data.xlsx")

    class_wb_creator(student_dict, x, sort)

#print(ws.max_row)
#print(ws.max_column)
def update_priority_ratings():
    wb = load_workbook("Student_Data.xlsx")
    ws = wb.active

    for row in range(2, ws.max_row + 1):
        future_test_date = ws["F" + str(row)].value # "F" is Future Test Date column
        check_testing(future_test_date)
        ws["D" + str(row)] = priority # "D" is Priority Rating column

    wb.save("Student_Data.xlsx")

def check_testing(future_test_date):
    global priority
    
    priority_rating_raw = dt.strptime(future_test_date, "%d %b, %Y") - dt.now() 
    priority_string = str(priority_rating_raw)
    priority_string.replace("days,", "    ")
    pr_string = priority_string.split(" ")[0]
    pr = int(pr_string[0:3]) #pr = priority_rating
    if pr > 61:
        priority = "D"
    elif 40< pr <=60:
        priority = "C"
    elif 20< pr <=40:
        priority = "B"
    elif 0< pr <=20:
        priority = "A"

# --------------Regular Expressions Legend--------------
# CTRL + F = Search and replace bar
# MetaCharacters (Need to be escaped):
# .[{()\^$?*+
# \. <---- "." has been "escaped" so that we look for the literal of "." and not everything

# .       - Any Character Except New Line
# \d      - Digit (0-9)
# \D      - Not a Digit (0-9)
# \w      - Word Charecter (a-z, A-Z, 0-9, _)
# \W      - Not a Word Charecter
# \s      - Whitespace (space, tab, newline)
# \S      - Not a Whitespace

# --------- Anchors ---------------
# \b      - Word Boundry
# \B      - Not a Word Boundry
# ^       - Beginning of a String
# $       - End of a string

# []      - Matches characters IN brackets
# [^ ]    - Matches charecters NOT in brackets
# |       - Either or
# ( )     - Group

# --------- Quantifiers -----------
# *       - 0 or more
# +       - 1 or more
# ?       - 0 or 1
# {3}     - Exact number
# {3,4}   - Range of numbers (Min, Max)    
    
def create_student():
    status = True
    first_name = input("Enter First Name: ")
    last_name = input("Enter Last Name: ")
    belt = input("Enter Belt Rank: ")

    last_test_date = input("Enter Previous Test Date (1 Jan, 2022). If N/A, press ENTER: ")
    if re.search(r"(\d{1,2}) (\w{3}), (\d{4})", last_test_date) == None:
        last_test_date = "N/A"

    future_test_date = input("Enter Next Test Date (1 Apr, 2022): ")
    if re.search(r"(\d{1,2}) (\w{3}), (\d{4})", future_test_date) == None:
        print("Incorrect Format. Try Again.")
        create_student()
    print("""
    ------- Class Legend ------ | --------- Format Eg's ---------
    Monday = Mon                |
    Tuesday = Tue               |            
    Wednsday = Wed              |            Mon-Adv
    Thursday = Thu              |
    Friday = Fri                |            Tue-Beg
    Saturday = Sat              |            
                                |            Wed-Int
    Beginner = Beg              |
    Intermediate = Int          |            Thu-Adu
    Advanced = Adv              |
    Teen = Tee                  |            Thu-Tee
    Adult = Adu                 |            
                                |            Fri-Int
    **************************  |            
    Int and Adv = InA           |            Sat-Beg
                                |            
    No Class = na               |            Sat-InA
    -----------------------------------------------------------        
    """)
    first_class = input("Enter 1st Class: ")
    if re.search(r"(\w{3})[-](\w{3})", first_class) == None:
        print("Incorrect Format. Try Again.")
    second_class = input("Enter 2nd Class: ")
    if second_class == "na":
        second_class = "N/A"
    elif re.search(r"(\w{3})[-](\w{3})", second_class) == None:
        print("Incorrect Format. Try Again.")
    
    check_testing(future_test_date)
    
    temp_dict = {
        first_name:{
            "Last Name" : last_name,
            "Belt" : belt,
            "Priority" : priority,
            "Last Test Date" : last_test_date,
            "Future Test Date" : future_test_date,
            "First Class" : first_class,
            "Second Class" : second_class,
            }
    }

    wb = load_workbook("Student_Data.xlsx")
    ws = wb.active

    for person in temp_dict:
        points = list(temp_dict[person].values())
        ws.append([person] + points)
    
    wb.save("Student_Data.xlsx")

    print("\n{} {} has been added to the data sheet.\n".format(first_name, last_name))
    display()

def display_student(row_value):
    wb = load_workbook("Student_Data.xlsx")
    ws = wb.active
    count = 1
    header_dict = {1 : "First Name: ", 2 : "Last Name: ", 3 : "Belt Rank: ", 4 : "Priority Rating: ", 5 : "Last Test Date: ", 6 : "Next Test Date: ", 7 : "First Class: ", 8 : "Second Class: "}

    student_row_index = list(ws.rows)[row_value]
  
    for data in student_row_index:
        if count == 9:
            break
        elif data.value == None:
            print(header_dict[count] + "N/A")
            count += 1
        else:
            print(header_dict[count] + data.value)
            count += 1
    wb.save("Student_Data.xlsx")
    display()
          
def view_student():
    wb = load_workbook("Student_Data.xlsx")
    ws = wb.active

    x = input("Input Name (First Last): ")
    f, l = x.split(" ")
    
    for row in range(2, ws.max_row + 1):
        first_name = ws["A" + str(row)].value
        if first_name == f:
            for row2 in range(2, ws.max_row + 1):
                    last_name_cell = ws["B" + str(row2)].value
                    if last_name_cell == l:
                        display_student(row - 1)
                        break
    wb.save("Student_Data.xlsx")

def display():
    print("""
    ------ Student Management ------
    
        vs = View Student Info
        vd = View the Day
        cs = Create Student Profile

    --------------------------------

    """)
    x = input("Input a Command: ")
    if x == "vs":
        view_student()
    elif x == "cs":
        create_student()
    elif x == "vd":
        wb_data_compiler()

update_priority_ratings()       
display()

